import openpyxl
from pathlib import Path
from natsort import natsorted

def combine_worksheet(folder):
    """Combine the worksheets in a folder."""
    # Check if not folder, print message and break.
    folder_path = Path(folder)
    if not folder_path.exists():
        print(f"Can't find folder {folder_path}")
        return
    
    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.title = 'Result of combine'
    
    current_row = 1
    
    # Get the files in folder,and get every file's worksheet and cell values.
    files = list(folder_path.glob('*.xlsx'))
    sorted_files = natsorted(files, key=lambda x: x.name)
    
    for file_path in sorted_files:
        
        print(f"Moving {file_path.name}")
        old_wb = openpyxl.load_workbook(file_path,data_only=True)
        old_sheet = old_wb.active
        for r in range(1, old_sheet.max_row+1):
            
            for c in range(1, old_sheet.max_column+1):
                old_value = old_sheet.cell(row=r, column=c).value

                # Move values to new worksheet.
                new_sheet.cell(row=current_row, column=c).value = old_value
            current_row += 1
        print(f"Finishing {file_path.name} move.")
    save_path = folder_path.parent/'Combine_result.xlsx'
    new_wb.save(save_path)

target_folder = '/Users/qfhxsxzw/test'
combine_worksheet(target_folder)